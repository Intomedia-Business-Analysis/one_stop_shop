# -*- coding: utf-8 -*-
"""Migrér historiske klippekort-data fra Excel-arket ind i KlippekortForbrug
+ Pipedrive-felterne 'klip brugt' og 'klip fra tidligere aftale'.

Excel sporer klip pr. KUNDE (samlet over flere aftaler). Hver kunde-fane har et
afviklet-klip-log i kolonnerne E-I (Klip-nr, Medie, Dato, Kontakt, Titel).

Model pr. matchet deal:
    total tilgængelige klip = clip_card_size (deal) + klip_fra_tidligere
    klip_fra_tidligere      = max(0, excel_total - clip_card_size)
    klip_brugt              = antal afviklede klip i Excel (rækker med dato)

To trin:
  1) python migrate_klippekort.py export   -> skriver klippekort_mapping.csv til gennemsyn/rettelse
  2) python migrate_klippekort.py import    -> dry-run (viser hvad der ville ske)
     python migrate_klippekort.py import --commit  -> udfører (DB-insert + Pipedrive-push)

'import' læser den (evt. rettede) klippekort_mapping.csv. Ret kolonnen
'matched_deal_id' for kunder uden sikkert match, før du kører import.
"""
import csv
import datetime as dt
import glob
import os
import re
import sys
import uuid

import openpyxl

from moduler.modul_klippekort import queries as q
from moduler.modul_klippekort.pipedrive_api import (
    JPPOL_TOKEN_ENV,
    USED_CLIP_FIELD_KEY,
    _get_token,
)

# Sti til Excel-arket: tag automatisk den NYESTE "Overvågning klippekort*.xlsm"
# i Downloads (så en ny download — fx "(1)" — bruges uden at ændre koden).
def _find_xlsx() -> str:
    pattern = os.path.expanduser(r"~\Downloads\Overvågning klippekort*.xlsm")
    matches = [p for p in glob.glob(pattern) if not os.path.basename(p).startswith("~$")]
    if not matches:
        return os.path.expanduser(r"~\Downloads\Overvågning klippekort.xlsm")
    return max(matches, key=os.path.getmtime)

XLSX = _find_xlsx()
MAPPING_CSV = os.path.join(os.path.dirname(os.path.abspath(__file__)), "klippekort_mapping.csv")

# Pipedrive custom-felt-key for "klip fra tidligere aftale" (oprettet af bruger)
PREV_CLIP_FIELD_KEY = "5103db826fb8b1580f694af93dc57152b2d6f674"

# Kilde-stempel på importerede rækker (så de kan genkendes/ryddes/idempotens)
IMPORT_TAG = "excel-import"

META_SHEETS = {"Navne på virksomheder", "Tidl. klippekortskunder", "Skabelon", "Skabelon (2)"}

# NB: " danmark"/" denmark" strippes IKKE — det indgår i rigtige firmanavne
# (fx "Sparekassen Danmark"), og stripping fik "Sparekassen Thy/Kronjylland"
# til fejlagtigt at matche "Sparekassen Danmark".
SUFFIXES = [" a/s", " as", " aps", " amba", " a.m.b.a", " holding",
            " group", " gruppen", " p/s", " k/s", " s/i", " i/s"]


def norm(s):
    if not s:
        return ""
    s = str(s).strip().lower()
    s = s.replace("æ", "ae").replace("ø", "oe").replace("å", "aa")
    s = re.sub(r"[^\w\s/+]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    for _ in range(3):
        for suf in SUFFIXES:
            if s.endswith(suf):
                s = s[: -len(suf)].strip()
    return s.replace(" ", "")


# ---------------------------------------------------------------------------
# Excel-læsning
# ---------------------------------------------------------------------------

def _as_date(v):
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    if isinstance(v, str):
        m = re.search(r"(\d{1,2})[-/.](\d{1,2})[-/.](\d{4})", v)
        if m:
            d, mo, y = (int(x) for x in m.groups())
            try:
                return dt.date(y, mo, d)
            except ValueError:
                return None
    return None


def read_sheet(ws):
    """Returnér dict med kunde-metadata + liste af afviklede klip."""
    head = list(ws.iter_rows(min_row=1, max_row=4, values_only=True))

    def cell(r, c):
        try:
            return head[r][c]
        except (IndexError, TypeError):
            return None

    a1 = cell(0, 0)
    udloeb = _as_date(cell(2, 6))           # G3
    total = cell(3, 4)                       # E4  (Antal klip)
    rest = cell(3, 5)                        # F4  (Klip tilbage)
    site_card = cell(3, 6)                   # G4  (medie/site for kortet)
    pris_total = cell(3, 7)                  # H4  (pris)
    try:
        total = int(total) if total is not None else None
    except (TypeError, ValueError):
        total = None
    try:
        rest = int(rest) if rest is not None else None
    except (TypeError, ValueError):
        rest = None
    try:
        pris_total = float(pris_total) if pris_total is not None else None
    except (TypeError, ValueError):
        pris_total = None

    all_rows = list(ws.iter_rows(min_row=1, values_only=True))

    used = []
    for row in all_rows[5:]:
        if len(row) < 9:
            continue
        medie, dato, titel = row[5], row[6], row[8]
        d = _as_date(dato)
        if d and (titel or medie):
            used.append({
                "site": str(medie or "").strip(),
                "stilling": str(titel or "").strip(),
                "tidspunkt": d.isoformat(),
            })

    # "Overvågning"-sektionen (venstre side): aktuelt opslåede stillinger med
    # 'Udløb WM' (slutdato) + evt. 'Effektgaranti'. Header genkendes på "Udløb WM".
    live = {}  # norm(stilling) -> {"slutdato": iso, "effektgaranti": bool}
    hdr = None
    for i, row in enumerate(all_rows):
        for c in range(0, 5):
            v = row[c] if c < len(row) else None
            if isinstance(v, str) and "udl" in v.lower() and "wm" in v.lower():
                hdr = i
                break
        if hdr is not None:
            break
    if hdr is not None:
        for row in all_rows[hdr + 1:]:
            stilling = row[0] if len(row) > 0 else None
            udl = _as_date(row[2]) if len(row) > 2 else None
            eff = row[3] if len(row) > 3 else None
            if stilling and udl:
                key = str(stilling).strip().lower()
                live[key] = {"slutdato": udl.isoformat(), "effektgaranti": bool(eff)}

    return {
        "sheet": ws.title, "a1": a1, "udloeb": udloeb, "total": total,
        "rest": rest, "site_card": site_card, "pris_total": pris_total,
        "used": used, "live": live,
    }


# ---------------------------------------------------------------------------
# Deal-matchning
# ---------------------------------------------------------------------------

def load_deals():
    conn = q.get_conn(); cur = conn.cursor(as_dict=True)
    cur.execute("""
        SELECT pd_deal_id, org_id, org_name,
               TRY_CAST(clip_card_size AS INT) AS size,
               CAST(value_dkk AS BIGINT) AS value_dkk,
               CONVERT(NVARCHAR(10), advertising_periode_slut, 23) AS slut,
               CASE WHEN CAST(GETDATE() AS date)
                         BETWEEN advertising_periode_start AND advertising_periode_slut
                    THEN 1 ELSE 0 END AS aktiv
        FROM [dbo].[PipedriveDeals]
        WHERE pipeline_name='job' AND account='jppol_advertising' AND status='won'
          AND TRY_CAST(clip_card_size AS INT) > 0
    """)
    rows = cur.fetchall(); conn.close()
    return rows


def build_index(deals):
    idx = {}
    for d in deals:
        idx.setdefault(norm(d["org_name"]), []).append(d)
    return idx


def candidates(by_norm, sheet_norm, a1_norm):
    out, seen = [], set()
    keys = {k for k in (sheet_norm, a1_norm) if k and len(k) >= 5}
    for dk, ds in by_norm.items():
        for k in keys:
            if dk == k or dk.startswith(k) or k.startswith(dk):
                for d in ds:
                    if d["pd_deal_id"] not in seen:
                        seen.add(d["pd_deal_id"]); out.append(d)
                break
    return out


def pick(cands, udloeb, total):
    """Vælg bedste deal + confidence ud fra udløbsdato og størrelse."""
    if not cands:
        return None, "none"
    def score(d):
        s = 0
        if udloeb and d["slut"] == udloeb.isoformat():
            s += 2
        if total and d["size"] == total:
            s += 1
        if d["aktiv"]:
            s += 1
        return s
    ranked = sorted(cands, key=lambda d: (score(d), d["slut"] or ""), reverse=True)
    best = ranked[0]; sc = score(best)
    conf = "high" if sc >= 2 else ("medium" if sc >= 1 else "low")
    if len(cands) == 1 and conf == "low":
        conf = "medium"
    return best, conf


# ---------------------------------------------------------------------------
# Mode: export mapping
# ---------------------------------------------------------------------------

def export():
    print(f"Kilde: {XLSX}")
    deals = load_deals()
    by_norm = build_index(deals)
    deals_by_id = {d["pd_deal_id"]: d for d in deals}

    # Bevar tidligere gennemgang: hvis mapping-filen findes, genbrug brugerens
    # matched_deal_id/note/confidence for faner der allerede er behandlet.
    prev = {}
    if os.path.exists(MAPPING_CSV):
        with open(MAPPING_CSV, "r", encoding="utf-8-sig") as f:
            for r in csv.DictReader(f, delimiter=";"):
                prev[r["sheet"]] = r
    n_preserved = n_new = 0

    wb = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)
    rows = []
    for ws in wb.worksheets:
        if ws.title in META_SHEETS:
            continue
        info = read_sheet(ws)
        total = info["total"]
        base = {
            "sheet": info["sheet"], "navn_a1": info["a1"] or "",
            "excel_udloeb": info["udloeb"].isoformat() if info["udloeb"] else "",
            "excel_total": total if total is not None else "",
            "excel_brugt": len(info["used"]),
            "excel_rest": info["rest"] if info["rest"] is not None else "",
        }
        p = prev.get(info["sheet"])
        if p is not None:
            # Bevar brugerens beslutning; opdatér kun data-felterne fra nyt ark.
            n_preserved += 1
            mid = str(p.get("matched_deal_id", "")).strip()
            row = {**base,
                   "matched_deal_id": mid,
                   "matched_org": p.get("matched_org", ""),
                   "deal_size": p.get("deal_size", ""),
                   "deal_slut": p.get("deal_slut", ""),
                   "deal_aktiv": p.get("deal_aktiv", ""),
                   "klip_fra_tidligere": "",
                   "n_deals_for_org": p.get("n_deals_for_org", ""),
                   "confidence": p.get("confidence", ""),
                   "note": p.get("note", "")}
            if mid.isdigit() and int(mid) in deals_by_id:
                d = deals_by_id[int(mid)]
                row["matched_org"] = d["org_name"]
                row["deal_size"] = d["size"]; row["deal_slut"] = d["slut"]; row["deal_aktiv"] = d["aktiv"]
                row["klip_fra_tidligere"] = max(0, (total or 0) - (d["size"] or 0))
            rows.append(row)
        else:
            # Ny fane i det opdaterede ark — kør matcher til gennemsyn.
            n_new += 1
            cands = candidates(by_norm, norm(info["sheet"]), norm(info["a1"]))
            deal, conf = pick(cands, info["udloeb"], total)
            if deal:
                rows.append({**base,
                    "matched_deal_id": deal["pd_deal_id"], "matched_org": deal["org_name"],
                    "deal_size": deal["size"], "deal_slut": deal["slut"], "deal_aktiv": deal["aktiv"],
                    "klip_fra_tidligere": max(0, (total or 0) - (deal["size"] or 0)),
                    "n_deals_for_org": len(cands), "confidence": conf, "note": "NY fane — tjek"})
            else:
                rows.append({**base,
                    "matched_deal_id": "", "matched_org": "", "deal_size": "",
                    "deal_slut": "", "deal_aktiv": "", "klip_fra_tidligere": "",
                    "n_deals_for_org": len(cands), "confidence": "none",
                    "note": "NY fane — UDFYLD matched_deal_id"})
    wb.close()

    # Faner der var i mapping men ikke længere i arket
    present = {r["sheet"] for r in rows}
    forsvundet = [s for s in prev if s not in present]

    cols = ["sheet", "navn_a1", "excel_udloeb", "excel_total", "excel_brugt",
            "excel_rest", "matched_deal_id", "matched_org", "deal_size",
            "deal_slut", "deal_aktiv", "klip_fra_tidligere", "n_deals_for_org",
            "confidence", "note"]
    with open(MAPPING_CSV, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=cols, delimiter=";")
        w.writeheader()
        w.writerows(rows)

    n_matched = sum(1 for r in rows if str(r["matched_deal_id"]).strip().isdigit())
    n_used = sum(int(r["excel_brugt"]) for r in rows)
    print(f"Skrev {len(rows)} kunder til {MAPPING_CSV}")
    print(f"  bevaret fra tidligere gennemgang: {n_preserved}, nye faner: {n_new}")
    print(f"  med deal_id (klar til import): {n_matched}")
    print(f"  afviklede klip i alt: {n_used}")
    if n_new:
        nye = [r["sheet"] for r in rows if "NY fane" in (r.get("note") or "")]
        print(f"  NYE faner at tjekke: {', '.join(nye)}")
    if forsvundet:
        print(f"  ADVARSEL: faner i mapping men IKKE i nyt ark: {', '.join(forsvundet)}")


# ---------------------------------------------------------------------------
# Mode: import (dry-run / commit)
# ---------------------------------------------------------------------------

def _read_mapping():
    if not os.path.exists(MAPPING_CSV):
        print(f"Mangler {MAPPING_CSV} — kør 'export' først.")
        sys.exit(1)
    with open(MAPPING_CSV, "r", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f, delimiter=";"))


def _import_already_done() -> int:
    """Antal allerede-importerede rækker i tabellen (created_by = IMPORT_TAG)."""
    try:
        conn = q.get_conn(); cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM KlippekortForbrug WHERE created_by=%s", (IMPORT_TAG,))
        n = int(cur.fetchone()[0]); conn.close()
        return n
    except Exception:
        return 0


def do_import(commit=False, force=False):
    from moduler.modul_klippekort.pipedrive_api import BASE_URL  # noqa
    import requests

    # Engangs-spærre: migreringen er en éngangs-backfill. Når den først er kørt,
    # er tabellen kilden — værktøjet kigger kun på KlippekortForbrug fremover.
    # En ny commit ville slette/overskrive de importerede rækker, så det kræver
    # eksplicit --force.
    if commit and not force:
        n = _import_already_done()
        if n > 0:
            print(f"STOP: migreringen er allerede kørt ({n} importerede rækker findes).")
            print("Fremover læser værktøjet blot fra tabellen. Brug --force hvis du")
            print("bevidst vil gen-importere (overskriver de importerede rækker).")
            return

    mapping = _read_mapping()
    deals = {d["pd_deal_id"]: d for d in load_deals()}
    wb = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)
    by_title = {ws.title: ws for ws in wb.worksheets}

    token = _get_token()
    if not token:
        print(f"ADVARSEL: {JPPOL_TOKEN_ENV} mangler i .env — Pipedrive-push springes over.")

    n_deals = n_rows = n_push = 0
    conn = q.get_conn(); cur = conn.cursor()
    for m in mapping:
        try:
            deal_id = int(m["matched_deal_id"])
        except (TypeError, ValueError):
            continue  # ikke matchet — spring over
        ws = by_title.get(m["sheet"])
        if ws is None:
            continue
        info = read_sheet(ws)
        deal = deals.get(deal_id)
        if not deal:
            print(f"  ! deal {deal_id} ({m['sheet']}) findes ikke i DB — springes over")
            continue

        excel_total = int(m["excel_total"]) if str(m.get("excel_total")).strip().isdigit() else (info["total"] or 0)
        size = deal["size"] or 0
        tidligere = max(0, excel_total - size)
        used = info["used"]
        live = info.get("live") or {}
        pris_pr_klip = round((info["pris_total"] / excel_total), 2) if (info["pris_total"] and excel_total) else (
            round(deal["value_dkk"] / size, 2) if size else 0.0)

        # Slutdato/effektgaranti fra "Overvågning"-sektionen sættes på den NYESTE
        # afvikling pr. stillingsbetegnelse (= den aktuelt opslåede), så vi ikke
        # sætter en fremtidig slutdato på gamle gentagelser af samme titel.
        latest_idx = {}
        for i, u in enumerate(used):
            key = u["stilling"].strip().lower()
            if key in live and (key not in latest_idx or u["tidspunkt"] > used[latest_idx[key]]["tidspunkt"]):
                latest_idx[key] = i
        n_slut = len(latest_idx)

        print(f"  {m['sheet']:35} -> deal {deal_id} {deal['org_name'][:32]:32} | "
              f"brugt={len(used)} tidligere={tidligere} pris/klip={pris_pr_klip} slutdatoer={n_slut}")
        n_rows += len(used)

        if commit:
            # Idempotens: fjern tidligere import-rækker for denne deal
            cur.execute("DELETE FROM KlippekortForbrug WHERE pd_deal_id=%s AND created_by=%s",
                        (deal_id, IMPORT_TAG))
            for i, u in enumerate(used):
                key = u["stilling"].strip().lower()
                slutdato, eff = None, 0
                if latest_idx.get(key) == i:
                    slutdato = live[key]["slutdato"]
                    eff = 1 if live[key]["effektgaranti"] else 0
                # Hver historisk afvikling = ét job på ét site, der koster 1 klip.
                cur.execute("""
                    INSERT INTO KlippekortForbrug
                        (job_id, pd_deal_id, org_id, org_name, site, stilling, tidspunkt,
                         slutdato, effektgaranti, klip, pris_pr_klip, clip_card_size, created_by)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (uuid.uuid4().hex, deal_id, deal["org_id"], deal["org_name"], u["site"],
                      u["stilling"], u["tidspunkt"], slutdato, eff, 1, pris_pr_klip, size, IMPORT_TAG))

            if token:
                payload = {USED_CLIP_FIELD_KEY: str(len(used)),
                           PREV_CLIP_FIELD_KEY: str(tidligere)}
                r = requests.put(f"{BASE_URL}/deals/{deal_id}",
                                 params={"api_token": token}, json=payload, timeout=30)
                if r.status_code < 400 and r.json().get("success"):
                    n_push += 1
                else:
                    print(f"    ! Pipedrive-push fejlede for {deal_id}: {r.status_code} {r.text[:160]}")
        n_deals += 1

    if commit:
        conn.commit()
    conn.close()
    wb.close()
    mode = "COMMIT" if commit else "DRY-RUN"
    print(f"\n[{mode}] kunder={n_deals} forbrug-rækker={n_rows} pipedrive-push={n_push}")
    if not commit:
        print("Kør med '--commit' for at udføre.")


if __name__ == "__main__":
    mode = sys.argv[1] if len(sys.argv) > 1 else "export"
    if mode == "export":
        export()
    elif mode == "import":
        do_import(commit="--commit" in sys.argv, force="--force" in sys.argv)
    else:
        print("Brug: python migrate_klippekort.py [export|import] [--commit] [--force]")
