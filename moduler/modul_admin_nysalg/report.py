"""Rapportgenerering: Excel (openpyxl) og PDF (reportlab).

Administrative nysalg kommer fra matchet (effective_is_admin, dvs. is_admin
justeret af direktørens override); administrative opsigelser hører til udtrækkets
`administrativ`-flag og indgår ikke her. Rapporten gemmes til report_dir og stien
returneres.
"""
from __future__ import annotations

import logging
import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from moduler.modul_admin_nysalg.repo import (effective_gross_in, effective_gross_out,
                                             effective_is_admin, is_admin_opsigelse)

logger = logging.getLogger(__name__)

# Engelsk, multi-valuta-rapport → ISO-koder (DKK/NOK/SEK/EUR) frem for "kr.",
# så DKK/NOK/SEK ikke forveksles.
_DKK = "#,##0 \"DKK\""


def _num_fmt(cur: str | None) -> str:
    """Excel-talformat med ISO-valuta-suffiks (fx '#,##0 "EUR"')."""
    return f'#,##0 "{cur or "DKK"}"'


# ── Land/type-gruppering (delt af Excel + PDF) ────────────────────────────────

def _row_metrics(b: dict) -> tuple[float, float, float, float, float]:
    """(actual_sale, actual_churn, net_growth, budget, deviation) for en brand-række.

    Alt ekskl. administrative bevægelser. net = sale − churn; deviation = net − budget.
    """
    sale = (b.get("brutto") or 0) - (b.get("adm_nysalg") or 0)
    churn = (b.get("opsigelser") or 0) - (b.get("adm_opsigelser") or 0)
    net = b.get("netto")
    net = (sale - churn) if net is None else net
    budget = b.get("budget") or 0
    return sale, churn, net, budget, net - budget


def _country_groups(brand_rows: list[dict] | None) -> list[dict]:
    """Gruppér brand-rækker → [{country, currency, types:[{type, rows}]}] i fast
    land-/type-rækkefølge (Subscription før Advertising). Ukendte brands havner i
    'Other'. Lande lægges ikke sammen — kun subtotaler pr. (land, type) og land."""
    from moduler.modul_admin_nysalg.brands import (COUNTRY_CURRENCY, COUNTRY_ORDER,
                                                   TYPE_ORDER, brand_geo)
    by: dict[str, dict[str, list]] = {}
    for b in brand_rows or []:
        country, typ = brand_geo(b.get("brand"))
        by.setdefault(country, {}).setdefault(typ, []).append(b)
    out: list[dict] = []
    for country in COUNTRY_ORDER + [c for c in by if c not in COUNTRY_ORDER]:
        tmap = by.get(country)
        if not tmap:
            continue
        types = []
        for typ in TYPE_ORDER + [t for t in tmap if t not in TYPE_ORDER]:
            rows = tmap.get(typ)
            if rows:
                types.append({"type": typ, "rows": rows})
        out.append({"country": country, "currency": COUNTRY_CURRENCY.get(country, "DKK"),
                    "types": types})
    return out


def _sum_metrics(rows: list[dict]) -> tuple[float, float, float, float, float]:
    """Summér _row_metrics over en gruppe rækker (deviation = Σnet − Σbudget)."""
    sale = churn = net = budget = 0.0
    for b in rows:
        s, c, n, bu, _ = _row_metrics(b)
        sale += s; churn += c; net += n; budget += bu
    return sale, churn, net, budget, net - budget


_HEAD_FILL = PatternFill("solid", fgColor="1C1C1A")  # hub-primær (nær-sort)
_HEAD_FONT = Font(bold=True, color="FFFFFF", size=10)
_TITLE_FONT = Font(bold=True, size=14)
_THIN = Side(style="thin", color="DDDDDD")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def report_dir() -> str:
    d = os.getenv("ADMIN_NYSALG_REPORT_DIR") or os.path.join("data", "admin_nysalg_reports")
    os.makedirs(d, exist_ok=True)
    return d


def _base_filename(run: dict) -> str:
    pf = (run.get("period_from") or "").strip()
    pt = (run.get("period_to") or "").strip()
    if pf or pt:
        period = f"{pf or 'start'}_{pt or 'slut'}"
    else:
        period = (run.get("period") or "alle").replace("/", "-")
    return f"monthly-performance-report-{run['run_id']}-{period}"


# ── Excel ────────────────────────────────────────────────────────────────────

def _style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = _HEAD_FILL
        cell.font = _HEAD_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = _BORDER


def _autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def generate_excel(run: dict, matches: list[dict], summary: dict,
                   brand_rows: list[dict] | None = None, out_dir: str | None = None,
                   pd_deals: list[dict] | None = None,
                   org_names: dict | None = None,
                   months_breakdown: list[dict] | None = None,
                   site_rows: list[dict] | None = None) -> str:
    out_dir = out_dir or report_dir()
    org_names = org_names or {}
    wb = Workbook()

    # ── Ark 1: Summary ────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Monthly performance report"
    ws["A1"].font = _TITLE_FONT
    meta = [
        ("Period", run.get("period") or "—"),
        ("Run ID", run.get("run_id")),
        ("Source file", run.get("source_filename") or run.get("source_path") or "—"),
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Approved by", run.get("approved_by") or "—"),
        ("Approved", run["approved_at"].strftime("%Y-%m-%d %H:%M")
            if run.get("approved_at") else "—"),
    ]
    r = 3
    for label, val in meta:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws.cell(row=r, column=2, value=val)
        r += 1

    # Pr. land → type med subtotaler. Alle tal ekskl. administrative bevægelser.
    # Lande lægges IKKE sammen (forskellige valutaer) — kun subtotal pr. (land,
    # type) og total pr. land. Beløb i lokal valuta (DKK/NOK/SEK/EUR).
    r += 1
    head_row = r
    bcols = ["Brand", "Actual Sale", "Actual Churn", "Actual Net Growth",
             "Budget Net Growth", "Deviation", "Comment"]
    for i, h in enumerate(bcols, start=1):
        ws.cell(row=head_row, column=i, value=h)
    _style_header(ws, head_row, len(bcols))
    r += 1

    _SECTION_FILL = PatternFill("solid", fgColor="EDEAE3")   # land-overskrift
    _SUBTOTAL_FILL = PatternFill("solid", fgColor="F7F5F2")   # subtotal/total

    def _vals(row, vals, cf, bold=False, fill=None):
        """Skriv kol 2–6 (sale, churn, net, budget, deviation) med format/stil."""
        for i, v in enumerate(vals, start=2):
            c = ws.cell(row=row, column=i, value=v)
            c.number_format = cf
            if bold:
                c.font = Font(bold=True)
            if fill:
                c.fill = fill

    for grp in _country_groups(brand_rows):
        cur = grp["currency"]
        cf = _num_fmt(cur)
        # Land-overskrift
        hc = ws.cell(row=r, column=1, value=f"{grp['country']} ({cur})")
        hc.font = Font(bold=True, size=11)
        for i in range(1, len(bcols) + 1):
            ws.cell(row=r, column=i).fill = _SECTION_FILL
        r += 1
        for tb in grp["types"]:
            ws.cell(row=r, column=1, value=f"  {tb['type']}").font = Font(bold=True, italic=True)
            r += 1
            for b in tb["rows"]:
                sale, churn, net, budget, dev = _row_metrics(b)
                ws.cell(row=r, column=1, value="    " + b["brand"])
                has_budget = b.get("budget") is not None
                _vals(r, [sale, churn, net,
                          budget if has_budget else None,
                          dev if has_budget else None], cf)
                ws.cell(row=r, column=7, value=b.get("comment") or "")
                r += 1
                for s in b.get("subrows") or []:
                    s_sale, s_churn, s_net, _, _ = _row_metrics(s)
                    ws.cell(row=r, column=1, value="      ↳ " + s["brand"])
                    _vals(r, [s_sale, s_churn, s_net, None, None], cf)
                    r += 1
            ssale, schurn, snet, sbudget, sdev = _sum_metrics(tb["rows"])
            ws.cell(row=r, column=1,
                    value=f"  Subtotal {tb['type']}").font = Font(bold=True)
            _vals(r, [ssale, schurn, snet, sbudget, sdev], cf, bold=True,
                  fill=_SUBTOTAL_FILL)
            ws.cell(row=r, column=1).fill = _SUBTOTAL_FILL
            ws.cell(row=r, column=7).fill = _SUBTOTAL_FILL
            r += 1
        # Total pr. land (på tværs af typer)
        all_rows = [b for tb in grp["types"] for b in tb["rows"]]
        tsale, tchurn, tnet, tbudget, tdev = _sum_metrics(all_rows)
        ws.cell(row=r, column=1, value=f"Total {grp['country']}").font = Font(bold=True, size=11)
        _vals(r, [tsale, tchurn, tnet, tbudget, tdev], cf, bold=True, fill=_SECTION_FILL)
        for i in (1, 7):
            ws.cell(row=r, column=i).fill = _SECTION_FILL
        r += 2

    _autosize(ws, [26, 15, 15, 18, 18, 15, 40])

    # Overall comment
    r += 1
    ws.cell(row=r, column=1, value="Overall comment").font = Font(bold=True)
    ws.cell(row=r + 1, column=1, value=run.get("director_comment") or "—")

    # ── Niches: WM DK + WM NO (pr. niche/site) ────────────────────────────────
    if site_rows:
        wsn = wb.create_sheet("Niches WM DK & WM NO")
        ncols = ["Niche", "Actual Sale", "Actual Churn", "Actual Net Growth"]
        wsn.append(ncols)
        _style_header(wsn, 1, len(ncols))
        from moduler.modul_admin_nysalg.brands import COUNTRY_CURRENCY
        cur_brand = {"Watch DK": None, "Watch NO": None}
        for g in site_rows:
            # Sektions-overskrift når brandet skifter (Watch DK / Watch NO).
            br = g["brand"]
            if cur_brand.get(br) is None:
                cur_brand[br] = True
                wsn.append([f"{br} ({g.get('currency') or 'DKK'})"])
                hr = wsn.max_row
                wsn.cell(row=hr, column=1).font = Font(bold=True, size=11)
                wsn.cell(row=hr, column=1).fill = PatternFill("solid", fgColor="EDEAE3")
            cf = _num_fmt(g.get("currency"))
            wsn.append(["    " + g["site"], g["sale"], g["churn"], g["net"]])
            row = wsn.max_row
            for col in (2, 3, 4):
                wsn.cell(row=row, column=col).number_format = cf
        _autosize(wsn, [30, 16, 16, 18])

    # ── Ark 2: Administrative nysalg (det der trækkes fra) ─────────────────────
    ws2 = wb.create_sheet("Administrative new sales")
    cols = ["Row", "Month", "Site", "Customer", "Org ID", "Account", "Movement",
            "Gross in", "Matched deal", "Deal value", "Pipeline", "Override", "Comment"]
    ws2.append(cols)
    _style_header(ws2, 1, len(cols))
    for m in matches:
        if not effective_is_admin(m):
            continue
        ws2.append([
            m.get("row_index"), m.get("month_end"), m.get("site"), m.get("matched_org_name"),
            m.get("pipedrive_id"), m.get("account_number"), m.get("movement"), m.get("gross_in"),
            m.get("matched_deal_id"), m.get("matched_value"), m.get("matched_pipeline"),
            m.get("override") or "", m.get("row_comment") or "",
        ])
    for row_cells in ws2.iter_rows(min_row=2, min_col=8, max_col=8):
        for c in row_cells:
            c.number_format = _DKK
    for row_cells in ws2.iter_rows(min_row=2, min_col=10, max_col=10):
        for c in row_cells:
            c.number_format = _DKK
    _autosize(ws2, [7, 11, 28, 28, 9, 14, 12, 14, 12, 14, 16, 10, 40])

    # ── Ark 3: Bevægelser pr. brand (Zuora) ───────────────────────────────────
    # Alle rækker med gross in eller gross out i perioden — kilden bag brand-
    # totalerne, så de kan holdes op imod PipeDrive-deals (Ark 4).
    ws4 = wb.create_sheet("Movements per brand")
    # Gross in/out = EFFEKTIVE værdier (direktørens medtag/udeluk + værdiretning),
    # så arket matcher totalerne. Excluded/Override viser hvor der er rettet.
    mcols = ["Brand", "Site", "Customer", "Org ID", "Account", "Month", "Movement",
             "Currency", "Gross in", "Gross out", "Net", "Excluded", "Override",
             "Administrative", "Matched deal"]
    ws4.append(mcols)
    _style_header(ws4, 1, len(mcols))
    moves = sorted(
        [m for m in matches if (m.get("gross_in") or 0) or (m.get("gross_out") or 0)],
        key=lambda m: ((m.get("brand") or ""), (m.get("site") or "")))
    from moduler.modul_admin_nysalg.brands import brand_account
    for m in moves:
        gi = effective_gross_in(m)
        go = effective_gross_out(m)
        # Navn fra matchet deal, ellers slå op på org-id i brandets KONTO (org-id
        # er ikke unikke på tværs af konti) — også for ikke-admin rækker.
        acct = brand_account(m.get("brand") or "")
        pid = str(m.get("pipedrive_id") or "").strip()
        kunde = (m.get("matched_org_name")
                 or (org_names.get(acct, {}).get(pid, "") if acct else ""))
        # Administrativ = matchede en administrativ deal (nysalg ELLER opsigelse)
        # eller bærer Zuoras administrativ-flag.
        adm = "Yes" if (effective_is_admin(m) or is_admin_opsigelse(m)) else ""
        excluded = "Yes" if m.get("total_excluded") else ""
        override = ("Yes" if (m.get("gross_in_override") is not None
                              or m.get("gross_out_override") is not None) else "")
        ws4.append([
            m.get("brand") or "", m.get("site"), kunde,
            m.get("pipedrive_id"), m.get("account_number"), m.get("month_end"),
            m.get("movement"), m.get("currency") or "", gi, go, gi - go,
            excluded, override, adm, m.get("matched_deal_id") or "",
        ])
    for col in (9, 10, 11):
        for cells in ws4.iter_rows(min_row=2, min_col=col, max_col=col):
            for c in cells:
                c.number_format = "#,##0"
    _autosize(ws4, [16, 26, 26, 9, 14, 11, 12, 8, 14, 14, 14, 9, 9, 12, 12])
    ws4.auto_filter.ref = ws4.dimensions   # filtrerbar Zuora-tabel

    # ── Ark 4: PipeDrive-deals (måned) ────────────────────────────────────────
    # Alle won-deals med service_activation_date i perioden — PipeDrive-kilden,
    # til afstemning mod Zuora-bevægelserne (Ark 3).
    ws5 = wb.create_sheet("PipeDrive deals")
    pcols = ["Brand", "Site", "Customer", "Org ID", "Account", "Team", "Pipeline",
             "Status", "Administrative", "Service act. date", "Currency", "Value"]
    ws5.append(pcols)
    _style_header(ws5, 1, len(pcols))
    for d in (pd_deals or []):
        ws5.append([
            d.get("brand") or "", d.get("site"), d.get("org_name"), d.get("org_id"),
            d.get("account"), d.get("team"), d.get("pipeline"), d.get("status"),
            "Yes" if d.get("administrativ") else "", d.get("service_activation_date"),
            d.get("currency") or "", d.get("value") or 0,
        ])
    for cells in ws5.iter_rows(min_row=2, min_col=12, max_col=12):
        for c in cells:
            c.number_format = "#,##0"
    _autosize(ws5, [16, 26, 26, 9, 18, 22, 16, 8, 12, 16, 8, 14])
    ws5.auto_filter.ref = ws5.dimensions   # filtrerbar PipeDrive-tabel

    # ── Ark: Performance pr. måned ────────────────────────────────────────────
    # Én række pr. (måned, brand) — så perioden kan brydes ned på de enkelte
    # måneder. Summen pr. brand over alle måneder matcher "Summary"-arket.
    if months_breakdown:
        ws6 = wb.create_sheet("Per month")
        mcols2 = ["Month", "Brand", "Actual Sale", "Actual Churn",
                  "Actual Net Growth", "Budget Net Growth"]
        ws6.append(mcols2)
        _style_header(ws6, 1, len(mcols2))
        for blk in months_breakdown:
            for b in blk.get("rows", []):
                cf = _num_fmt(b.get("currency"))   # lokal valuta (NO/SE/DE)
                sale, churn, net, budget, _ = _row_metrics(b)
                ws6.append([blk["label"], b["brand"], sale, churn, net,
                            budget if b.get("budget") is not None else None])
                row = ws6.max_row
                for col in (3, 4, 5, 6):
                    ws6.cell(row=row, column=col).number_format = cf
        ws6.auto_filter.ref = ws6.dimensions
        _autosize(ws6, [16, 22, 15, 15, 18, 18])

    path = os.path.join(out_dir, _base_filename(run) + ".xlsx")
    wb.save(path)
    return path


# ── PDF ──────────────────────────────────────────────────────────────────────

def generate_pdf(run: dict, matches: list[dict], summary: dict,
                 brand_rows: list[dict] | None = None, out_dir: str | None = None,
                 months_breakdown: list[dict] | None = None) -> str:
    out_dir = out_dir or report_dir()
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer, Table,
                                        TableStyle)
    except ImportError as e:
        raise RuntimeError("reportlab er ikke installeret — PDF kan ikke genereres.") from e

    # ── Palet ──────────────────────────────────────────────────────────────────
    # Hub-palet (static/hub.css): JP|Politiken nær-sort + varm beige.
    DARK    = colors.HexColor("#1C1C1A")   # --green (primær brand, nær-sort)
    INK     = colors.HexColor("#1A1A17")   # --ink
    MUTED   = colors.HexColor("#8B857A")   # --ink-3
    FAINT   = colors.HexColor("#BEB7AC")   # --ink-4 (banner-subtitle på mørk)
    LIGHT   = colors.HexColor("#F7F5F2")   # --bg (alternerende rækker)
    BORDER  = colors.HexColor("#E9E5DE")   # --border
    WIN     = colors.HexColor("#1F8A5B")   # --win (positiv netto)
    RED     = colors.HexColor("#B32E45")   # --danger (negativ)
    TAUPE   = colors.HexColor("#ACA199")   # --taupe (JP|Politiken signatur-bar)
    WHITE   = colors.white

    def money(v, cur="DKK"):
        n = f"{round(v or 0):,.0f}".replace(",", ".")
        return f"{n} {cur or 'DKK'}"   # ISO-kode (DKK/NOK/SEK/EUR)

    # ── Stilarter ────────────────────────────────────────────────────────────
    s_title   = ParagraphStyle("t",  fontName="Helvetica-Bold", fontSize=21, textColor=WHITE, leading=23)
    s_sub     = ParagraphStyle("s",  fontName="Helvetica",      fontSize=9,  textColor=FAINT, leading=13)
    s_logo    = ParagraphStyle("lg", fontName="Helvetica-Bold", fontSize=12, textColor=WHITE, leading=13)
    s_logosub = ParagraphStyle("ls", fontName="Helvetica-Bold", fontSize=7,  textColor=FAINT, leading=9)
    s_kpi_lbl = ParagraphStyle("kl", fontName="Helvetica-Bold", fontSize=7,  textColor=MUTED, leading=10, spaceAfter=3)
    s_kpi_val = ParagraphStyle("kv", fontName="Helvetica-Bold", fontSize=15, textColor=INK,   leading=17)
    s_kpi_sub = ParagraphStyle("ks", fontName="Helvetica",      fontSize=6.5, textColor=MUTED, leading=9, spaceBefore=2)
    s_h       = ParagraphStyle("h",  fontName="Helvetica-Bold", fontSize=11, textColor=INK,   leading=14, spaceBefore=10, spaceAfter=8)
    s_body    = ParagraphStyle("b",  fontName="Helvetica",      fontSize=9,  textColor=INK,   leading=13)
    s_cell    = ParagraphStyle("c",  fontName="Helvetica",      fontSize=8,  textColor=INK,   leading=10)
    s_cellb   = ParagraphStyle("cb", fontName="Helvetica-Bold", fontSize=8,  textColor=INK,   leading=10)
    s_hd      = ParagraphStyle("hd", fontName="Helvetica-Bold", fontSize=7,  textColor=WHITE, leading=9)

    CW = 178 * mm  # indholdsbredde (A4 − margener)

    def _footer(canvas, doc_):
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(MUTED)
        canvas.drawString(16 * mm, 10 * mm,
                          f"Sales & Marketing JP Politiken Business Media · Monthly performance report · generated {datetime.now():%Y-%m-%d %H:%M}")
        canvas.drawRightString(194 * mm, 10 * mm, f"Page {doc_.page}")
        canvas.setStrokeColor(BORDER)
        canvas.setLineWidth(0.5)
        canvas.line(16 * mm, 13 * mm, 194 * mm, 13 * mm)
        canvas.restoreState()

    path = os.path.join(out_dir, _base_filename(run) + ".pdf")
    doc = SimpleDocTemplate(path, pagesize=A4, leftMargin=16 * mm, rightMargin=16 * mm,
                            topMargin=15 * mm, bottomMargin=18 * mm,
                            title="Monthly performance report")
    el = []

    # ── Header-banner med JP|Politiken-logo ─────────────────────────────────────
    meta = (f"Period {run.get('period') or 'all'}  ·  Run #{run.get('run_id')}  ·  "
            f"Approved by {run.get('approved_by') or '—'}")
    # Logoet er hubbens typografiske wordmark: taupe signatur-bar + JP|Politiken.
    logo_bar = Table([[""]], colWidths=[20 * mm], rowHeights=[1.8 * mm])
    logo_bar.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), TAUPE),
        ("LEFTPADDING", (0, 0), (-1, -1), 0), ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0), ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))
    banner = Table([[[
        logo_bar,
        Spacer(1, 4),
        Paragraph("JP|Politiken", s_logo),
        Paragraph("BUSINESS MEDIA", s_logosub),
        Spacer(1, 11),
        Paragraph("Monthly performance report", s_title),
        Paragraph(meta, s_sub),
    ]]], colWidths=[CW])
    banner.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), DARK),
        ("LINEBELOW", (0, 0), (-1, -1), 3, TAUPE),
        ("LEFTPADDING", (0, 0), (-1, -1), 20), ("RIGHTPADDING", (0, 0), (-1, -1), 20),
        ("TOPPADDING", (0, 0), (-1, -1), 18), ("BOTTOMPADDING", (0, 0), (-1, -1), 18),
    ]))
    el.append(banner)
    el.append(Spacer(1, 9 * mm))

    # ── KPI-kort: net growth pr. land (lokal valuta, ingen grand total) ─────────
    groups = _country_groups(brand_rows)
    def kpi_cell(label, value, sub, color):
        return [Paragraph(label, s_kpi_lbl),
                Paragraph(value, ParagraphStyle("kvx", parent=s_kpi_val, textColor=color)),
                Paragraph(sub, s_kpi_sub)]
    if groups:
        cards = []
        for grp in groups:
            all_rows = [b for tb in grp["types"] for b in tb["rows"]]
            _, _, net, budget, dev = _sum_metrics(all_rows)
            cards.append(kpi_cell(
                f"{grp['country'].upper()} · NET GROWTH",
                money(net, grp["currency"]),
                f"budget {money(budget, grp['currency'])} · dev {money(dev, grp['currency'])}",
                WIN if net >= 0 else RED))
        n = len(cards)
        kpis = Table([cards], colWidths=[CW / n] * n)
        kstyle = [
            ("LINEABOVE", (0, 0), (-1, 0), 2.5, DARK),
            ("BACKGROUND", (0, 0), (-1, -1), WHITE),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 10), ("RIGHTPADDING", (0, 0), (-1, -1), 10),
            ("TOPPADDING", (0, 0), (-1, -1), 12), ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
        ]
        for i in range(n):
            kstyle.append(("BOX", (i, 0), (i, 0), 0.75, BORDER))
        kpis.setStyle(TableStyle(kstyle))
        el.append(kpis)
        el.append(Spacer(1, 9 * mm))

    # ── Performance pr. land og type ───────────────────────────────────────────
    # Grupperet land → type (Subscription/Advertising) med subtotaler og land-total.
    # Alle tal ekskl. administrative bevægelser, i lokal valuta. NET = Actual Net
    # Growth, farvet grøn/rød; DEV = afvigelse til budget (net − budget).
    el.append(Paragraph("Performance by country &amp; type", s_h))
    # Egen Currency-kolonne pr. række (DKK/NOK/SEK/EUR); tallene står uden valuta-
    # endelse, så tabellen er ren. NET/DEV-indeks er efter Brand + Currency.
    NET, DEV = 4, 6

    def num(v):
        return "" if v is None else f"{round(v or 0):,.0f}".replace(",", ".")

    head = ["", "Cur.", "Actual Sale", "Actual Churn", "Actual Net Growth",
            "Budget Net Growth", "Deviation"]
    data = [[Paragraph(h, s_hd) for h in head]]
    country_idx, type_idx, sub_idx, total_idx = [], [], [], []
    color_cells = []   # (col, rowidx, positive)

    def _row(label_para, vals, cur):
        return [label_para, cur] + [num(v) for v in vals]

    for grp in groups:
        cur = grp["currency"]
        ci = len(data)
        country_idx.append(ci)
        data.append([Paragraph(f"{grp['country']} ({cur})",
                               ParagraphStyle("ch", parent=s_hd, fontSize=8))] + [""] * 6)
        for tb in grp["types"]:
            type_idx.append(len(data))
            data.append([Paragraph(tb["type"],
                                   ParagraphStyle("ty", parent=s_cellb, textColor=MUTED))]
                        + [""] * 6)
            for b in tb["rows"]:
                sale, churn, net, budget, dev = _row_metrics(b)
                has_b = b.get("budget") is not None
                ri = len(data)
                data.append(_row(Paragraph(b["brand"], s_cell),
                                 [sale, churn, net,
                                  budget if has_b else None, dev if has_b else None], cur))
                color_cells.append((NET, ri, net >= 0))
                if has_b:
                    color_cells.append((DEV, ri, dev >= 0))
                for s in b.get("subrows") or []:
                    s_sale, s_churn, s_net, _, _ = _row_metrics(s)
                    sri = len(data)
                    data.append(_row(Paragraph("↳ " + s["brand"],
                                               ParagraphStyle("sr", parent=s_cell, textColor=MUTED)),
                                     [s_sale, s_churn, s_net, None, None], cur))
                    color_cells.append((NET, sri, s_net >= 0))
            ssale, schurn, snet, sbudget, sdev = _sum_metrics(tb["rows"])
            sub_idx.append(len(data))
            data.append(_row(Paragraph(f"Subtotal {tb['type']}", s_cellb),
                             [ssale, schurn, snet, sbudget, sdev], cur))
            color_cells.append((NET, sub_idx[-1], snet >= 0))
            color_cells.append((DEV, sub_idx[-1], sdev >= 0))
        all_rows = [b for tb in grp["types"] for b in tb["rows"]]
        tsale, tchurn, tnet, tbudget, tdev = _sum_metrics(all_rows)
        total_idx.append(len(data))
        data.append(_row(Paragraph(f"Total {grp['country']}", s_cellb),
                         [tsale, tchurn, tnet, tbudget, tdev], cur))
        color_cells.append((NET, total_idx[-1], tnet >= 0))
        color_cells.append((DEV, total_idx[-1], tdev >= 0))

    if len(data) > 1:
        t = Table(data, colWidths=[34 * mm, 13 * mm, 26 * mm, 26 * mm, 26 * mm,
                                   26 * mm, 27 * mm], repeatRows=1)
        style = [
            ("BACKGROUND", (0, 0), (-1, 0), DARK),
            ("TOPPADDING", (0, 0), (-1, 0), 7), ("BOTTOMPADDING", (0, 0), (-1, 0), 7),
            ("FONTNAME", (1, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (1, 1), (-1, -1), 7),
            ("TEXTCOLOR", (1, 1), (-1, -1), INK),
            ("TEXTCOLOR", (1, 1), (1, -1), MUTED),   # Currency-kolonne nedtonet
            ("ALIGN", (2, 0), (-1, -1), "RIGHT"),    # tal højrestillet
            ("ALIGN", (1, 0), (1, -1), "CENTER"),    # Currency centreret
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LINEBELOW", (0, 1), (-1, -1), 0.4, BORDER),
            ("LEFTPADDING", (0, 0), (-1, -1), 5), ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 1), (-1, -1), 5), ("BOTTOMPADDING", (0, 1), (-1, -1), 5),
        ]
        for ci in country_idx:
            style += [("BACKGROUND", (0, ci), (-1, ci), DARK),
                      ("TEXTCOLOR", (0, ci), (-1, ci), WHITE),
                      ("SPAN", (0, ci), (-1, ci)),
                      ("TOPPADDING", (0, ci), (-1, ci), 6),
                      ("BOTTOMPADDING", (0, ci), (-1, ci), 6)]
        for ti in type_idx:
            style.append(("BACKGROUND", (0, ti), (-1, ti), LIGHT))
            style.append(("SPAN", (0, ti), (-1, ti)))
        for si in sub_idx:
            style += [("BACKGROUND", (0, si), (-1, si), LIGHT),
                      ("FONTNAME", (1, si), (-1, si), "Helvetica-Bold"),
                      ("LINEABOVE", (0, si), (-1, si), 0.4, MUTED)]
        for to in total_idx:
            style += [("BACKGROUND", (0, to), (-1, to), colors.HexColor("#EDEAE3")),
                      ("FONTNAME", (0, to), (-1, to), "Helvetica-Bold"),
                      ("LINEABOVE", (0, to), (-1, to), 0.8, DARK)]
        for col, ri, pos in color_cells:
            style.append(("TEXTCOLOR", (col, ri), (col, ri), WIN if pos else RED))
            style.append(("FONTNAME", (col, ri), (col, ri), "Helvetica-Bold"))
        t.setStyle(TableStyle(style))
        el.append(t)
        el.append(Spacer(1, 12 * mm))

    # ── Net growth pr. måned (kun Denmark/DKK) ─────────────────────────────────
    # Bryder perioden ned på de enkelte måneder. Lande har forskellige valutaer og
    # lægges ikke sammen — tabellen viser derfor kun Denmark (DKK). NO/SE/DE-detaljer
    # pr. måned ses i Excel-fanen "Per month".
    from moduler.modul_admin_nysalg.brands import brand_geo as _brand_geo
    if months_breakdown and len(months_breakdown) > 1:
        el.append(Paragraph("Net growth per month — Denmark (DKK)", s_h))
        mhead = ["Month", "Actual Sale", "Actual Churn", "Actual Net Growth"]
        mdata = [[Paragraph(h, s_hd) for h in mhead]]
        msigns = []
        for blk in months_breakdown:
            rows = [b for b in blk.get("rows", [])
                    if _brand_geo(b.get("brand"))[0] == "Denmark"]
            sale = sum((b["brutto"] - b["adm_nysalg"]) for b in rows)
            churn = sum((b["opsigelser"] - b["adm_opsigelser"]) for b in rows)
            net = sale - churn
            msigns.append(net >= 0)
            mdata.append([Paragraph(blk["label"], s_cellb), money(sale, "DKK"),
                          money(churn, "DKK"), money(net, "DKK")])
        mt = Table(mdata, colWidths=[44 * mm, 45 * mm, 45 * mm, 44 * mm], repeatRows=1)
        mstyle = [
            ("BACKGROUND", (0, 0), (-1, 0), DARK),
            ("TOPPADDING", (0, 0), (-1, 0), 7), ("BOTTOMPADDING", (0, 0), (-1, 0), 7),
            ("FONTNAME", (1, 1), (-1, -1), "Helvetica"), ("FONTSIZE", (1, 1), (-1, -1), 8),
            ("TEXTCOLOR", (1, 1), (-1, -1), INK), ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [WHITE, LIGHT]),
            ("LINEBELOW", (0, 1), (-1, -1), 0.4, BORDER),
            ("LEFTPADDING", (0, 0), (-1, -1), 7), ("RIGHTPADDING", (0, 0), (-1, -1), 7),
            ("TOPPADDING", (0, 1), (-1, -1), 6), ("BOTTOMPADDING", (0, 1), (-1, -1), 6),
        ]
        for i, pos in enumerate(msigns, start=1):
            mstyle.append(("TEXTCOLOR", (3, i), (3, i), WIN if pos else RED))
            mstyle.append(("FONTNAME", (3, i), (3, i), "Helvetica-Bold"))
        mt.setStyle(TableStyle(mstyle))
        el.append(mt)
        el.append(Spacer(1, 12 * mm))

    # ── Kommentarer pr. brand ──────────────────────────────────────────────────
    brand_comments = [b for b in (brand_rows or []) if (b.get("comment") or "").strip()]
    if brand_comments:
        el.append(Paragraph("Comments per brand", s_h))
        for b in brand_comments:
            el.append(Paragraph(f"<b>{b['brand']}:</b> {b['comment'].replace(chr(10), '<br/>')}", s_body))
            el.append(Spacer(1, 2 * mm))
        el.append(Spacer(1, 6 * mm))

    # ── Samlede kommentar ──────────────────────────────────────────────────────
    if run.get("director_comment"):
        el.append(Paragraph("Overall comment", s_h))
        box = Table([[Paragraph(run["director_comment"].replace("\n", "<br/>"), s_body)]],
                    colWidths=[CW])
        box.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), LIGHT),
            ("LINEABOVE", (0, 0), (-1, -1), 2, DARK),
            ("LEFTPADDING", (0, 0), (-1, -1), 12), ("RIGHTPADDING", (0, 0), (-1, -1), 12),
            ("TOPPADDING", (0, 0), (-1, -1), 10), ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ]))
        el.append(box)

    doc.build(el, onFirstPage=_footer, onLaterPages=_footer)
    return path


def generate_report(run: dict, matches: list[dict], summary: dict,
                    brand_rows: list[dict] | None = None, fmt: str = "xlsx",
                    out_dir: str | None = None, pd_deals: list[dict] | None = None,
                    org_names: dict | None = None,
                    months_breakdown: list[dict] | None = None,
                    site_rows: list[dict] | None = None) -> str:
    """fmt: 'xlsx' | 'pdf'. Returnerer stien til den genererede fil."""
    if fmt == "pdf":
        return generate_pdf(run, matches, summary, brand_rows, out_dir,
                            months_breakdown=months_breakdown)
    return generate_excel(run, matches, summary, brand_rows, out_dir, pd_deals, org_names,
                          months_breakdown=months_breakdown, site_rows=site_rows)
