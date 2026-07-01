import io
import logging
import threading
from datetime import datetime

from fastapi import APIRouter, Body, Depends, HTTPException, Request
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from fastapi.templating import Jinja2Templates

from auth import get_current_user, has_access
from moduler.modul_klippekort.queries import (
    db_forbrug_for_deal,
    db_oekonomi,
    db_org_owner_meta,
    db_overblik,
    db_rediger_job,
    db_registrer_forbrug,
    db_slet_job,
    db_udloebende_jobs,
    get_site_groups,
    init_klippekort_db,
    missing_org_ids,
    refresh_missing_org_owners,
    refresh_org_owners,
)
from moduler.modul_klippekort.pipedrive_api import add_used_clip_cards

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/tools/klippekort", tags=["Klippekort"])
templates = Jinja2Templates(directory="templates")
from nav_utils import register_nav_globals
register_nav_globals(templates)
init_klippekort_db()

# Org-ejer cache opdateres i baggrunden (org-ejer er ikke et deal-felt i
# PipedriveDeals). Trigges ved sidevisning hvis cachen er tom/forældet.
_ORG_REFRESH_LOCK = threading.Lock()
_ORG_REFRESHING = {"running": False}
_ORG_STALE_HOURS = 24


def _maybe_refresh_pd_cache():
    # Fuld opdatering når cachen er tom eller forældet (>24t). Ellers tjek om
    # der er dukket NYE kunder op (needed org'er som endnu ikke er cachet) —
    # fx en deal hvis annonceringsperiode netop er startet — og hent kun dem.
    # Uden det sidste ville en ny kunde stå uden ejer indtil hele cachen udløb.
    meta = db_org_owner_meta()
    stale = meta["count"] == 0 or meta["alder_timer"] >= _ORG_STALE_HOURS
    missing = [] if stale else missing_org_ids()
    if not stale and not missing:
        return
    with _ORG_REFRESH_LOCK:
        if _ORG_REFRESHING["running"]:
            return
        _ORG_REFRESHING["running"] = True

    def _worker():
        try:
            if stale:
                refresh_org_owners()          # fuld opdatering (alle org'er)
            else:
                refresh_missing_org_owners()  # kun nyligt tilkomne kunder
        except Exception:
            logger.exception("Baggrunds-refresh af org-ejere fejlede")
        finally:
            _ORG_REFRESHING["running"] = False

    threading.Thread(target=_worker, daemon=True).start()


@router.get("/", response_class=HTMLResponse)
async def klippekort_page(request: Request, user=Depends(get_current_user)):
    if not has_access(user, "salesperson"):
        raise HTTPException(403, "Ingen adgang")
    return templates.TemplateResponse(request, "klippekort_overblik.html", {
        "user": user,
    })


@router.get("/overblik")
async def klippekort_overblik(mine: int = 0, status: str = "aktive", user=Depends(get_current_user)):
    if not has_access(user, "salesperson"):
        raise HTTPException(403, "Ingen adgang")
    if status not in ("aktive", "udloebne"):
        status = "aktive"
    _maybe_refresh_pd_cache()
    try:
        owner = user.get("name") if mine else None
        return JSONResponse({"rows": db_overblik(owner, status)})
    except Exception:
        logger.exception("klippekort_overblik fejlede (mine=%s, status=%s)", mine, status)
        raise HTTPException(500, "Data kunne ikke hentes")


@router.get("/export")
async def klippekort_export(user=Depends(get_current_user)):
    """Excel-fil over ALLE aktive klippekort til Koncern Økonomi ved månedsluk:
    oprindelige klip, brugte klip og resterende klip pr. kunde/klippekort."""
    if not has_access(user, "salesperson"):
        raise HTTPException(403, "Ingen adgang")
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        rows = db_overblik(None, "aktive")
        wb = Workbook()
        ws = wb.active
        ws.title = "Aktive klippekort"
        cols = ["Kunde", "Sælger", "Medie(r)", "Stilling", "Deal-ID",
                "Oprindelige klip", "Brugte klip", "Resterende klip",
                "Værdi (DKK)", "Periode start", "Periode slut", "Dage til udløb"]
        ws.append(cols)
        head_fill = PatternFill("solid", fgColor="1C1C1A")
        head_font = Font(bold=True, color="FFFFFF", size=10)
        for c in range(1, len(cols) + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = head_fill
            cell.font = head_font
            cell.alignment = Alignment(horizontal="left", vertical="center")
        for r in rows:
            ws.append([
                r.get("org_name"), r.get("owner_name"), r.get("sites"), r.get("title"),
                r.get("pd_deal_id"), r.get("clip_card_size"), r.get("brugt"), r.get("rest"),
                r.get("value_dkk"), r.get("periode_start"), r.get("periode_slut"),
                r.get("dage_til_udloeb"),
            ])
        for i, w in enumerate([34, 22, 28, 30, 10, 16, 13, 16, 14, 13, 13, 14], start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f"aktive-klippekort-{datetime.now():%Y-%m}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'},
        )
    except Exception:
        logger.exception("klippekort_export fejlede")
        raise HTTPException(500, "Eksporten kunne ikke genereres")


@router.get("/udloebende")
async def klippekort_udloebende(mine: int = 0, status: str = "aktive", user=Depends(get_current_user)):
    """Stillinger med slutdato — opfølgnings-oversigt (aktive eller udløbne)."""
    if not has_access(user, "salesperson"):
        raise HTTPException(403, "Ingen adgang")
    if status not in ("aktive", "udloebne"):
        status = "aktive"
    try:
        owner = user.get("name") if mine else None
        return JSONResponse({"rows": db_udloebende_jobs(owner, status)})
    except Exception:
        logger.exception("klippekort_udloebende fejlede (mine=%s, status=%s)", mine, status)
        raise HTTPException(500, "Data kunne ikke hentes")


@router.get("/sites")
async def klippekort_sites(user=Depends(get_current_user)):
    """Site-familier (Watch DK / Monitor) til 'opret job'-dropdownen."""
    if not has_access(user, "salesperson"):
        raise HTTPException(403, "Ingen adgang")
    return JSONResponse(get_site_groups())


@router.get("/oekonomi")
async def klippekort_oekonomi(user=Depends(get_current_user)):
    if not has_access(user, "salesperson"):
        raise HTTPException(403, "Ingen adgang")
    try:
        return JSONResponse({"rows": db_oekonomi()})
    except Exception:
        logger.exception("klippekort_oekonomi fejlede")
        raise HTTPException(500, "Data kunne ikke hentes")


@router.get("/forbrug")
async def klippekort_forbrug(pd_deal_id: int, user=Depends(get_current_user)):
    """Hent alle registrerede jobs/klip-forbrug for én deal."""
    if not has_access(user, "salesperson"):
        raise HTTPException(403, "Ingen adgang")
    try:
        return JSONResponse({"rows": db_forbrug_for_deal(pd_deal_id)})
    except Exception:
        logger.exception("klippekort_forbrug fejlede (pd_deal_id=%s)", pd_deal_id)
        raise HTTPException(500, "Data kunne ikke hentes")


@router.post("/slet-job")
async def klippekort_slet_job(payload: dict = Body(...), user=Depends(get_current_user)):
    """Slet et helt job (alle dets sites) og opdatér 'klip brugt' i Pipedrive.

    Body: {job_id}
    """
    if not has_access(user, "salesperson"):
        raise HTTPException(403, "Ingen adgang")
    job_id = (payload.get("job_id") or "").strip()
    if not job_id:
        raise HTTPException(400, "job_id er påkrævet")

    res = db_slet_job(job_id)
    if not res.get("ok"):
        raise HTTPException(400, res.get("error", "Kunne ikke slette job"))

    # Additivt: træk de fjernede klip fra Pipedrives autoritative used_clip_cards
    # (delta er negativ), så vi ikke overskriver klip toolet ikke selv har registreret.
    pd = add_used_clip_cards(res["pd_deal_id"], res["delta"])
    return JSONResponse({"ok": True, "brugt": res["brugt"], "pipedrive": pd})


@router.post("/rediger-job")
async def klippekort_rediger_job(payload: dict = Body(...), user=Depends(get_current_user)):
    """Redigér et aktivt stillingsopslag: forlæng perioden og/eller giv effektgaranti.

    Effektgaranti = aftalt med kunden hvor lang tid ekstra stillingen kører —
    derfor skal der altid en (ny) slutdato med.
    Body: {job_id, slutdato (YYYY-MM-DD), effektgaranti}
    """
    if not has_access(user, "salesperson"):
        raise HTTPException(403, "Ingen adgang")
    job_id = (payload.get("job_id") or "").strip()
    slutdato = (payload.get("slutdato") or "").strip()
    effektgaranti = bool(payload.get("effektgaranti"))
    if not job_id:
        raise HTTPException(400, "job_id er påkrævet")
    if not slutdato:
        raise HTTPException(400, "slutdato er påkrævet")

    res = db_rediger_job(job_id, slutdato, effektgaranti)
    if not res.get("ok"):
        raise HTTPException(400, res.get("error", "Kunne ikke opdatere job"))
    return JSONResponse({"ok": True})


@router.post("/opret-job")
async def klippekort_opret_job(payload: dict = Body(...), user=Depends(get_current_user)):
    """Registrér ét job (kan dække flere sites og koste flere klip) + push til Pipedrive.

    Body: {pd_deal_id, sites: [..], stilling, tidspunkt (YYYY-MM-DD), klip}
    """
    if not has_access(user, "salesperson"):
        raise HTTPException(403, "Ingen adgang")

    try:
        pd_deal_id = int(payload.get("pd_deal_id"))
    except (TypeError, ValueError):
        raise HTTPException(400, "pd_deal_id skal være et heltal")
    sites = payload.get("sites")
    if isinstance(sites, str):
        sites = [sites]
    sites = [s.strip() for s in (sites or []) if s and s.strip()]
    stilling  = (payload.get("stilling") or "").strip()
    tidspunkt = (payload.get("tidspunkt") or "").strip()
    slutdato  = (payload.get("slutdato") or "").strip() or None
    effektgaranti = bool(payload.get("effektgaranti"))
    try:
        klip = int(payload.get("klip", 1))
    except (TypeError, ValueError):
        klip = 1
    if not sites or not stilling or not tidspunkt:
        raise HTTPException(400, "mindst ét site, stilling og tidspunkt er påkrævet")
    if klip < 1:
        raise HTTPException(400, "antal klip skal være mindst 1")

    try:
        res = db_registrer_forbrug(pd_deal_id, sites, stilling, tidspunkt, klip,
                                   user.get("name") or "system", slutdato=slutdato,
                                   effektgaranti=effektgaranti)
        if not res.get("ok"):
            raise HTTPException(400, res.get("error", "Kunne ikke registrere job"))

        # Additivt: læg de nyligt registrerede klip oveni Pipedrives autoritative
        # used_clip_cards (delta = +klip) i stedet for at overskrive med toolets sum.
        pd = add_used_clip_cards(pd_deal_id, res["delta"])
        return JSONResponse({
            "ok":        True,
            "brugt":     res["brugt"],
            "rest":      res["rest"],
            "pipedrive": pd,
        })
    except HTTPException:
        raise
    except Exception:
        logger.exception("klippekort_opret_job fejlede (pd_deal_id=%s)", pd_deal_id)
        raise HTTPException(500, "Data kunne ikke hentes")
